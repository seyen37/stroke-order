"""
Build 教育部本土語言成果參考字表 (6,792 字) cover-set dataset.

CANONICAL SOURCE: Official xlsx published by 教育部國民及學前教育署
(113 年 1 月 2 日 臺教國署國字第 1120183030 號函).

Earlier name: 「臺灣本土語言用字參考字表」
Current name: 「教育部本土語言成果參考字表」(2024 renaming)

Distinguishing features vs 教育部 4808:
- Includes Hokkien/Hakka/Indigenous-language-specific characters
- Each entry has CNS 11643 codepoint (Taiwan one-true-source for char encoding)
- Companion 附表 (598 chars) tracks OS font support
  (Microsoft 新細明/正黑、Google 思源、Apple 蘋方)

Output: src/stroke_order/components/coversets/bentu_6792.json

Each entry includes:
- moe_index: 編號 1..6792
- unicode: derived from actual char codepoint
- cns11643: official Taiwan encoding (e.g. "1-4421")
- big5: legacy Taiwan encoding (when in Big5 range)
- stroke_count, radical: from xlsx columns
- os_support: dict (only for chars in 附表) of font support flags

Usage:
    python3 scripts/build_bentu_6792.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import opencc
from openpyxl import load_workbook


def _is_cjk_or_pua(c: str) -> bool:
    """Accept any single CJK ideograph OR PUA codepoint.

    PUA acceptance is intentional for this dataset: 教育部本土語言字表
    uses Private Use Area for Hokkien/Hakka/Indigenous-language-specific
    characters that don't have standard Unicode codepoints yet.
    """
    if len(c) != 1:
        return False
    code = ord(c)
    return (
        0x3400 <= code <= 0x4DBF        # CJK Ext A
        or 0x4E00 <= code <= 0x9FFF     # CJK Unified
        or 0x20000 <= code <= 0x3FFFF   # CJK Ext B-H
        or 0xE000 <= code <= 0xF8FF     # BMP PUA
        or 0xF0000 <= code <= 0xFFFFD   # Supplementary PUA-A
        or 0x100000 <= code <= 0x10FFFD # Supplementary PUA-B
    )


def parse_total_sheet(ws) -> list:
    """Parse 總表. Skip 2 header rows. Cols: 編號, 字元, 楷體, 宋體,
    Unicode, 區段, CNS碼位, Big5碼位, 筆畫, 部首."""
    out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:           # skip header rows
            continue
        if row[0] is None:  # empty / spacer row
            continue
        moe_idx = row[0]
        char = row[1]
        if not char or not _is_cjk_or_pua(str(char)):
            continue
        out.append({
            "moe_index":   int(moe_idx),
            "char":        str(char),
            "unicode":     row[4],          # may be "U+4E00" str
            "region":      row[5],          # URO / Ext-A / etc.
            "cns11643":    row[6],          # "1-4421" Taiwan canonical
            "big5":        row[7],          # "A440" or None
            "stroke_count": row[8],
            "radical":     row[9],
        })
    return out


def parse_appendix_sheet(ws) -> dict:
    """Parse 附表 → dict mapping moe_index → os_support dict.

    Cols: 編號, 字元, 楷體, 宋體, Unicode, 區段, CNS碼位, 筆畫, 部首,
          MS新細明, MS正黑, Google思源, Apple蘋方, 總表編號
    """
    out = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:           # skip header rows (3 levels of headers)
            continue
        if row[0] is None:
            continue
        # 總表編號 is in last col (index 13)
        total_idx = row[13]
        if total_idx is None:
            continue
        out[int(total_idx)] = {
            "ms_mingti":   row[9]   == "支援",
            "ms_zhenghei": row[10]  == "支援",
            "google_siyuan": row[11] == "支援",
            "apple_pingfang": row[12] == "支援",
        }
    return out


def main():
    src = Path("/sessions/friendly-dreamy-noether/mnt/uploads/"
               "教育部本土語言成果參考字表.xlsx")
    if not src.exists():
        print(f"ERROR: {src} not found")
        sys.exit(1)

    print(f"Loading {src} ...")
    wb = load_workbook(src)
    print(f"  Sheets: {wb.sheetnames}")

    total = parse_total_sheet(wb[wb.sheetnames[0]])
    print(f"  總表 entries: {len(total)}")

    appendix = parse_appendix_sheet(wb[wb.sheetnames[1]])
    print(f"  附表 entries (with OS support): {len(appendix)}")

    # Spot-check moe_index continuity
    indices = [e["moe_index"] for e in total]
    if indices != list(range(1, len(total) + 1)):
        gaps = [i for i in range(1, len(total) + 1) if i not in set(indices)]
        print(f"  WARNING: moe_index gaps: {gaps[:10]}...")

    # Generate simp counterparts (TW → CN)
    t2s = opencc.OpenCC("t2s")
    chars_simp = [t2s.convert(e["char"]) for e in total]
    simp_diff = sum(1 for s, e in zip(chars_simp, total) if s != e["char"])
    print(f"  simp/trad differ: {simp_diff} ({100*simp_diff/len(total):.1f}%)")

    # Build entries
    entries = []
    for raw, simp in zip(total, chars_simp):
        entry = {
            "index":        raw["moe_index"],
            "moe_index":    raw["moe_index"],
            "trad":         raw["char"],
            "simp":         simp,
            "same":         simp == raw["char"],
            "unicode":      raw["unicode"] or f"U+{ord(raw['char']):04X}",
            "cns11643":     raw["cns11643"],
        }
        # Optional fields (omit when None to keep JSON tidy)
        if raw["big5"]:
            entry["big5"] = raw["big5"]
        if raw["stroke_count"] is not None:
            entry["stroke_count"] = raw["stroke_count"]
        if raw["radical"]:
            entry["radical"] = raw["radical"]
        # OS support (only present for chars in 附表)
        if raw["moe_index"] in appendix:
            entry["os_support"] = appendix[raw["moe_index"]]
        entries.append(entry)

    out = {
        "title": "教育部本土語言成果參考字表 6792",
        "english_title":
            "MOE-Taiwan Local Languages Reference Character Set (6,792)",
        "source":
            "中華民國教育部國民及學前教育署《教育部本土語言成果參考字表》"
            "（113 年 1 月 2 日 臺教國署國字第 1120183030 號函；"
            "原名「臺灣本土語言用字參考字表」）",
        "url": "https://language.moe.gov.tw/",
        "description":
            "台灣教育部官方公告之本土語言（國語、台語、客語、原住民語）"
            "用字參考表，共 6,792 字。包含標準漢字 + 本土語言特有字，"
            "每字含 CNS 11643 碼位 (Taiwan 一手公文最權威編碼)。"
            "598 個冷僻字（多為本土語言特有）附 OS 字型支援表 "
            "(Microsoft / Google / Apple)。",
        "char_count": len(entries),
        "trad_simp_mismatch_count": simp_diff,
        "appendix_count": len(appendix),
        "entries": entries,
    }

    out_path = Path(
        "src/stroke_order/components/coversets/bentu_6792.json"
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\nWrote {out_path}: {len(entries)} chars")
    print(f"  with CNS 11643 codepoints + {len(appendix)} OS-support entries")


if __name__ == "__main__":
    main()
